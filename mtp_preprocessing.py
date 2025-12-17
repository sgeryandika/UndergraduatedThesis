"""
mtp_preprocessing.py

Preprocessing for MPD -> cleaned CSV/XLSX + optimizer_input.csv (in ./output/)

- Adds optional progress_callback(current_row, total_rows) for UI integration.
- Creates 'Categories' from Check_Categories column (Line -> 1, Base/ Hangar -> 0; default 1).
- Exports:
    ./output/<output_csv>             (full cleaned CSV)
    ./output/optimizer_input.csv      (filtered subset for optimizer)
Returns: (df_final, optimizer_path, total_rows, num_valid_optimizer_rows)

Author: sgeryandika
"""

import re
from pathlib import Path
import argparse
import json

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def clean_mpd(file_path,
              output_csv="task_occurrence(2).csv",
              FH_per_day=8,
              FH_per_fc=1.5,
              FH_per_mo=None,
              output_dir="output",
              progress_callback=None,
              progress_batch=100):
    """
    Clean MPD and write outputs to output_dir.

    progress_callback(current_row:int, total_rows:int) -> None
        Called approximately every `progress_batch` rows (and at start/end) if provided.
    """
    if FH_per_mo is None:
        FH_per_mo = 30 * FH_per_day

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Read first sheet of the workbook
    xls = pd.ExcelFile(file_path)
    df_raw = xls.parse(0)
    total_rows = len(df_raw)

    # Inform callback of start
    if progress_callback:
        try:
            progress_callback(0, total_rows)
        except Exception:
            pass

    # --- Rename columns to normalized names (apply only if present) ---
    rename_map = {
        "TASK\nNUMBER": "TaskID",
        "ZONE": "Zone",
        "SKILL CODE": "SkillCode",
        "TASK CODE": "TaskCode",
        "TASK\nM.H.": "TaskMH",
        "ACCESS\nM.H.": "AccessMH",
        "PREP.\nM.H.": "PrepMH",
        "MEN": "Men",
        "TCI": "TimeControlledItem",
        "TPS": "TPS",
        "SAMPLE\nTHRESHOLD": "SampleThreshold",
        "SAMPLE\nINTERVAL": "SampleInterval",
        "100%\nTHRESHOLD": "Threshold100",
        "100%\nINTERVAL": "Interval100",
        # sometimes columns are plain 'THRESHOLD'/'INTERVAL'
        "THRESHOLD": "Threshold100",
        "INTERVAL": "Interval100",
    }
    df = df_raw.rename(columns={k: v for k, v in rename_map.items() if k in df_raw.columns}).copy()

    # --- Threshold parser helper ---
    def parse_threshold_with_flags(val):
        if pd.isna(val) or str(val).strip() == "":
            return {"fh": 0, "note": 0, "act": 0, "ce": 0}
        v = str(val).upper().replace("\n", " ")
        note_flag = 1 if "NOTE" in v else 0
        act_flag = 1 if "ACT" in v else 0
        ce_flag = 1 if "CE" in v else 0
        matches = re.findall(r"(\d+)\s*(YE|MO|DY|FC|FH|ACT YE)", v)
        fh_values = []
        for num_s, unit in matches:
            num = int(num_s)
            if unit in ("YE", "ACT YE"):
                fh_values.append(num * 365 * FH_per_day)
            elif unit == "MO":
                fh_values.append(num * FH_per_mo)
            elif unit == "DY":
                fh_values.append(num * FH_per_day)
            elif unit == "FC":
                fh_values.append(num * FH_per_fc)
            elif unit == "FH":
                fh_values.append(num)
        fh_val = int(min(fh_values)) if fh_values else 0
        return {"fh": fh_val, "note": note_flag, "act": act_flag, "ce": ce_flag}

    threshold_cols_all = ["SampleThreshold", "SampleInterval", "Threshold100", "Interval100"]
    threshold_cols = [c for c in threshold_cols_all if c in df.columns]

    # --- Parse rows with batched progress updates ---
    # We'll compute threshold-derived columns per-row and place them into df.
    for i, row in df.iterrows():
        for col in threshold_cols:
            parsed = parse_threshold_with_flags(row[col]) if col in df.columns else {"fh": 0, "note": 0, "act": 0, "ce": 0}
            df.at[i, f"{col}_FH"] = parsed["fh"]
            df.at[i, f"{col}_NOTE"] = int(parsed["note"])
            df.at[i, f"{col}_ACT"] = int(parsed["act"])
            df.at[i, f"{col}_CE"] = int(parsed["ce"])

        # batched progress callback
        if progress_callback and (i % progress_batch == 0 or i == total_rows - 1):
            try:
                progress_callback(i + 1, total_rows)
            except Exception:
                pass

    # --- Fallback logic ---
    if "SampleThreshold_FH" in df.columns and "Threshold100_FH" in df.columns:
        df["SampleThreshold_FH"] = np.where(df["SampleThreshold_FH"] == 0, df["Threshold100_FH"], df["SampleThreshold_FH"])
    if "SampleInterval_FH" in df.columns and "Interval100_FH" in df.columns:
        df["SampleInterval_FH"] = np.where(df["SampleInterval_FH"] == 0, df["Interval100_FH"], df["SampleInterval_FH"])

    # --- Numeric conversions ---
    for col in ["TaskMH", "AccessMH", "PrepMH", "Men"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

     # --- Fill missing values and conversions (do only if cols exist) ---
    if "TimeControlledItem" in df.columns:
        # accept 'T' => 1, otherwise 0. Also handle numeric 1/0
        def conv_tcu(x):
            if pd.isna(x):
                return 0
            sx = str(x).strip().upper()
            if sx == "T":
                return 1
            try:
                return int(float(sx))
            except:
                return 0
        df["TimeControlledItem"] = df["TimeControlledItem"].apply(conv_tcu)

    # --- TPS parsing ---
    if "TPS" in df.columns:
        def parse_tps(v):
            if pd.isna(v) or str(v).strip() == "":
                return {"TPS1": 0, "TPS2": 0, "TPS3": 0}
            s = str(v).strip().upper()
            if s == "NO":
                return {"TPS1": 0, "TPS2": 0, "TPS3": 0}
            return {"TPS1": 1 if "1" in s else 0, "TPS2": 1 if "2" in s else 0, "TPS3": 1 if "3" in s else 0}
        tps_flags = df["TPS"].apply(parse_tps)
        df["TPS1"] = tps_flags.apply(lambda x: x["TPS1"])
        df["TPS2"] = tps_flags.apply(lambda x: x["TPS2"])
        df["TPS3"] = tps_flags.apply(lambda x: x["TPS3"])

    # --- ManhoursRT derivation ---
    # --- Identify and normalize possible manhour columns ---
    mh_aliases = {
        "taskmh": "TaskMH",
        "task mh": "TaskMH",
        "accessmh": "AccessMH",
        "access mh": "AccessMH",
        "prep.mh": "PrepMH",
        "prep mh": "PrepMH",
        "manhours": "Manhours",
        "man hours": "Manhours",
        "mh": "Manhours"
    }

    # Normalize column names (case-insensitive match)
    df.columns = [c.strip() for c in df.columns]
    for col in df.columns:
        key = col.strip().lower().replace("\n", " ").replace("_", " ")
        if key in mh_aliases and mh_aliases[key] not in df.columns:
            df.rename(columns={col: mh_aliases[key]}, inplace=True)

    # Ensure numeric conversion for all MH-like columns
    mh_cols = [c for c in df.columns if any(tag in c.lower() for tag in ["mh", "manhour", "man hour"])]
    if not mh_cols:
        # fallback if no MH column at all
        df["ManhoursRT"] = 0
    else:
        for c in mh_cols:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
        df["ManhoursRT"] = df[mh_cols].sum(axis=1)

    # --- MenZero flag ---
    if "Men" in df.columns:
        df["MenZero"] = df["Men"].apply(lambda x: 1 if (pd.isna(x) or float(x) == 0.0) else 0)
    else:
        df["Men"] = 0
        df["MenZero"] = 1

    # --- Categories flag from Check_Categories ---
    def detect_categories(row):
        # Default: Line
        cat_val = 1

        # Check the Check_Categories column
        if "Check_Categories" in row and pd.notna(row["Check_Categories"]):
            s = str(row["Check_Categories"]).upper()
            if "BASE" in s or "HANGAR" in s:
                cat_val = 0
            elif "LINE" in s:
                cat_val = 1

        # Task Code rules
        if "Task Code" in row and pd.notna(row["Task Code"]):
            tc = str(row["Task Code"]).upper()
            if "C-CHECK" in tc or "DOCK" in tc or "SHOP" in tc:
                cat_val = 0

        # Skill Code rules
        if "Skill Code" in row and pd.notna(row["Skill Code"]):
            sc = str(row["Skill Code"]).upper()
            if sc == "NDT":
                # Look deeper into Description
                if "Description" in row and pd.notna(row["Description"]):
                    desc = str(row["Description"]).upper()
                    if "ULTRASONIC" in desc or "OVERHAUL" in desc:
                        cat_val = 0

        # Description rules
        if "Description" in row and pd.notna(row["Description"]):
            desc = str(row["Description"]).upper()
            if "MAJOR OVERHAUL" in desc or "STRUCTURAL REPAIR" in desc:
                cat_val = 0

        return cat_val

    df["Categories"] = df.apply(detect_categories, axis=1)
    
    # --- RowID sequential integers ---
    df.insert(0, "RowID", range(1, len(df) + 1))

    # --- Create IntervalFH canonical column (prefer FH-parsed columns) ---
    if "Interval100_FH" in df.columns:
        df["IntervalFH"] = pd.to_numeric(df["Interval100_FH"], errors="coerce").fillna(0).astype(float)
    elif "SampleInterval_FH" in df.columns:
        df["IntervalFH"] = pd.to_numeric(df["SampleInterval_FH"], errors="coerce").fillna(0).astype(float)
    else:
        # try raw Interval100 or SampleInterval numeric values
        if "Interval100" in df.columns:
            df["IntervalFH"] = pd.to_numeric(df["Interval100"], errors="coerce").fillna(0).astype(float)
        elif "SampleInterval" in df.columns:
            df["IntervalFH"] = pd.to_numeric(df["SampleInterval"], errors="coerce").fillna(0).astype(float)
        else:
            df["IntervalFH"] = 0.0

    # --- Final full cleaned dataframe ---
    df_final = df.copy()

    # write cleaned CSV
    out_csv_path = out_dir / output_csv
    df_final.to_csv(out_csv_path, index=False)

    # --- Build optimizer_input.csv (filtered) ---
    df_optimizer = df_final[(df_final["IntervalFH"] > 0) & (df_final["ManhoursRT"] > 0)].copy()
    cols_for_optimizer = [c for c in ["RowID", "TaskID", "IntervalFH", "ManhoursRT", "Men", "Categories", "Zone"] if c in df_optimizer.columns]
    df_opt = df_optimizer[cols_for_optimizer]
    optimizer_path = out_dir / "optimizer_input.csv"
    df_opt.to_csv(optimizer_path, index=False)

    # --- NEW: Base/Hangar Tasks List (Categories == 0) ---
    hangar_cols = [
        "RowID", "TaskID", "Description", "Skill Code", "Task Code",
        "IntervalFH", "Justification"
    ]
    # Keep only those columns if they exist in df_final
    hangar_cols_existing = [c for c in hangar_cols if c in df_final.columns]

    base_hangar_df = df_final[df_final["Categories"] == 0][hangar_cols_existing].copy()
    base_hangar_path = out_dir / "Base_Hangar_Tasks_List.csv"
    base_hangar_df.to_csv(base_hangar_path, index=False)


    # Final callback to indicate completion
    if progress_callback:
        try:
            progress_callback(total_rows, total_rows)
        except Exception:
            pass

    print(f"[INFO] Cleaned CSV saved to: {out_csv_path}")
    print(f"[INFO] Optimizer input saved to: {optimizer_path}")
    print(f"[INFO] Base/Hangar Tasks List saved to: {base_hangar_path}")

    return df_final, optimizer_path, total_rows, len(df_opt)


def export_to_excel(df, output_path):
    """Export dataframe to an Excel workbook with basic formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned MPD Data"
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        for cell in ws[r_idx]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4F81BD")
    # auto-width
    for col in ws.columns:
        max_length = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    wb.save(output_path)
    print(f"[INFO] Excel exported to: {output_path}")


# CLI convenience
def load_conversion_file(path: Path):
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    fh_day = float(data.get("FH_per_day", 8))
    fh_fc = float(data.get("FH_per_fc", 1.5))
    fh_mo = data.get("FH_per_mo")
    if fh_mo is None:
        fh_mo = 30 * fh_day
    else:
        fh_mo = float(fh_mo)
    return fh_day, fh_fc, fh_mo


def build_arg_parser():
    p = argparse.ArgumentParser(description="MPD preprocessing (cleaning + export)")
    p.add_argument("-i", "--input", required=True, help="Input MPD Excel file (first sheet used)")
    p.add_argument("-o", "--output-csv", default="task_occurrence(2).csv", help="Output CSV path")
    p.add_argument("--output-xlsx", default=None, help="Optional Excel export path")
    p.add_argument("--fh-per-day", type=float, default=None, help="Flight hours per day (overrides conv-file or default)")
    p.add_argument("--fh-per-fc", type=float, default=None, help="Flight hours per flight-cycle (overrides conv-file or default)")
    p.add_argument("--fh-per-mo", type=float, default=None, help="Flight hours per month (overrides conv-file or default)")
    p.add_argument("--conv-file", type=str, default=None, help="Optional JSON file supplying conversion constants")
    p.add_argument("--no-xlsx", action="store_true", help="Do not export to Excel even if --output-xlsx provided")
    return p


def main():
    parser = build_arg_parser()
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    FH_per_day = 8.0
    FH_per_fc = 1.5
    FH_per_mo = None

    if args.conv_file:
        fh_d, fh_fc, fh_mo = load_conversion_file(Path(args.conv_file))
        FH_per_day, FH_per_fc, FH_per_mo = fh_d, fh_fc, fh_mo

    if args.fh_per_day is not None:
        FH_per_day = args.fh_per_day
    if args.fh_per_fc is not None:
        FH_per_fc = args.fh_per_fc
    if args.fh_per_mo is not None:
        FH_per_mo = args.fh_per_mo

    print(f"Using conversion constants: FH_per_day={FH_per_day}, FH_per_fc={FH_per_fc}, FH_per_mo={FH_per_mo}")

    df = clean_mpd(str(input_path), output_csv=args.output_csv, FH_per_day=FH_per_day, FH_per_fc=FH_per_fc, FH_per_mo=FH_per_mo)
    print(f"Saved cleaned CSV to {args.output_csv}")

    if args.output_xlsx and (not args.no_xlsx):
        export_to_excel(df, output_path=args.output_xlsx)
        print(f"Saved Excel to {args.output_xlsx}")


if __name__ == "__main__":
    main()
